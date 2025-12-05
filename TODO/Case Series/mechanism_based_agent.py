"""
Mechanism-Based Drug Repurposing Agent
Identifies all drugs with a given mechanism and analyzes repurposing opportunities
"""

from drug_repurposing_agent import DrugRepurposingAgent
import logging
from typing import List, Dict
import time

logger = logging.getLogger(__name__)


class MechanismBasedAgent(DrugRepurposingAgent):
    """Extended agent that searches by mechanism to find multiple drugs"""
    
    def analyze_mechanism(self, mechanism: str) -> Dict:
        """
        Analyzes all drugs with a given mechanism
        
        Args:
            mechanism: e.g., "CGRP receptor antagonist", "JAK inhibitor"
            
        Returns:
            Dict with drugs and their repurposing opportunities
        """
        logger.info(f"Starting mechanism-based analysis: {mechanism}")
        
        # Step 1: Find all drugs with this mechanism
        drugs = self._find_drugs_by_mechanism(mechanism)
        logger.info(f"Found {len(drugs)} drugs with mechanism: {mechanism}")
        
        # Step 2: Analyze each drug
        all_results = {
            'mechanism': mechanism,
            'drugs_analyzed': [],
            'metadata': {
                'total_opportunities': 0,
                'analysis_date': None
            }
        }
        
        for drug in drugs:
            logger.info(f"Analyzing drug: {drug}")
            try:
                result = self.analyze_drug(drug)
                all_results['drugs_analyzed'].append(result)
                all_results['metadata']['total_opportunities'] += len(result['case_series'])
                time.sleep(1)  # Rate limiting between drugs
            except Exception as e:
                logger.error(f"Error analyzing {drug}: {e}")
                continue
        
        return all_results
    
    def _find_drugs_by_mechanism(self, mechanism: str) -> List[str]:
        """Finds all approved drugs with a given mechanism"""
        
        # Search for drugs with this mechanism
        search_results = self._tavily_search(
            f"{mechanism} approved drugs FDA",
            max_results=10
        )
        
        prompt = f"""Based on these search results, identify ALL FDA-approved drugs that work via this mechanism: {mechanism}

Search Results:
{search_results}

Return ONLY a JSON array of drug names (generic names, not brand names).

Example format:
["rimegepant", "ubrogepant", "atogepant"]

CRITICAL: Return ONLY the JSON array, nothing else. No markdown, no explanation."""

        response = self.anthropic_client.messages.create(
            model=self.model,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        self._track_tokens(response.usage)
        
        try:
            content = response.content[0].text.strip()
            if content.startswith("```"):
                content = content.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            drugs = eval(content)  # Safe here as we control the prompt
            return drugs
        except Exception as e:
            logger.error(f"Error parsing drugs: {e}")
            return []
    
    def export_mechanism_analysis(self, results: Dict, output_path: str):
        """Exports mechanism-based analysis to Excel with multiple sheets"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Summary sheet
        summary = wb.active
        summary.title = "Summary"
        summary.column_dimensions['A'].width = 30
        summary.column_dimensions['B'].width = 50
        
        # Summary data
        summary_data = [
            ("Mechanism Analyzed", results['mechanism']),
            ("Total Drugs", len(results['drugs_analyzed'])),
            ("Total Opportunities", results['metadata']['total_opportunities']),
            ("", ""),
            ("Drug", "Opportunities Found")
        ]
        
        for idx, (key, value) in enumerate(summary_data, 1):
            summary.cell(row=idx, column=1).value = key
            summary.cell(row=idx, column=2).value = str(value)
            if idx <= 3 or idx == 5:
                summary.cell(row=idx, column=1).font = Font(bold=True)
                summary.cell(row=idx, column=2).font = Font(bold=True)
        
        row_idx = 6
        for drug_result in results['drugs_analyzed']:
            summary.cell(row=row_idx, column=1).value = drug_result['drug_name']
            summary.cell(row=row_idx, column=2).value = len(drug_result['case_series'])
            row_idx += 1
        
        # Create sheet for each drug
        for drug_result in results['drugs_analyzed']:
            if not drug_result['case_series']:
                continue
                
            sheet_name = drug_result['drug_name'][:31]  # Excel sheet name limit
            drug_sheet = wb.create_sheet(sheet_name)
            
            # Headers
            headers = [
                "Disease", "N", "Evidence Level", "Response Rate",
                "Effect Size", "Efficacy Summary", "Safety Summary",
                "US Prevalence", "Unmet Need", "Clinical Score",
                "Evidence Score", "Market Score", "Overall Priority"
            ]
            
            for col_idx, header in enumerate(headers, 1):
                cell = drug_sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                drug_sheet.column_dimensions[chr(64+col_idx)].width = 15
            
            # Data
            for row_idx, case in enumerate(drug_result['case_series'], 2):
                scores = case.get('scores', {})
                market = case.get('market_data', {})
                
                row_data = [
                    case.get('disease', ''),
                    case.get('n', ''),
                    case.get('evidence_level', ''),
                    case.get('response_rate', ''),
                    case.get('effect_size_description', ''),
                    case.get('efficacy_summary', ''),
                    case.get('safety_summary', ''),
                    market.get('us_prevalence_estimate', ''),
                    market.get('unmet_need_description', ''),
                    scores.get('clinical_signal', ''),
                    scores.get('evidence_quality', ''),
                    scores.get('market_opportunity', ''),
                    scores.get('overall_priority', '')
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    drug_sheet.cell(row=row_idx, column=col_idx).value = value
                    drug_sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True)
                
                drug_sheet.row_dimensions[row_idx].height = 60
        
        wb.save(output_path)
        logger.info(f"Mechanism analysis exported to {output_path}")


def main():
    """Example usage for mechanism-based search"""
    import os
    from datetime import datetime
    
    ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
    TAVILY_API_KEY = os.getenv("TAVILY_API_KEY")
    
    if not ANTHROPIC_API_KEY or not TAVILY_API_KEY:
        print("Error: Please set ANTHROPIC_API_KEY and TAVILY_API_KEY environment variables")
        return
    
    # Initialize agent
    agent = MechanismBasedAgent(
        anthropic_api_key=ANTHROPIC_API_KEY,
        tavily_api_key=TAVILY_API_KEY
    )
    
    # Analyze mechanism
    mechanism = "CGRP receptor antagonist"
    print(f"Analyzing mechanism: {mechanism}...")
    
    results = agent.analyze_mechanism(mechanism)
    
    # Export results
    output_file = f"mechanism_analysis_{mechanism.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    agent.export_mechanism_analysis(results, output_file)
    
    print(f"\nAnalysis complete!")
    print(f"Analyzed {len(results['drugs_analyzed'])} drugs")
    print(f"Found {results['metadata']['total_opportunities']} total repurposing opportunities")
    print(f"Total cost: ${(agent.total_input_tokens * 0.003 / 1000 + agent.total_output_tokens * 0.015 / 1000):.2f}")
    print(f"Results saved to: {output_file}")


if __name__ == "__main__":
    main()
