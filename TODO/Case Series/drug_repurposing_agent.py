"""
Drug Repurposing Opportunity Agent
Systematically identifies and analyzes case series for drug repurposing opportunities
"""

import anthropic
import requests
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import time
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DrugRepurposingAgent:
    """Main orchestrator for drug repurposing opportunity identification"""
    
    def __init__(self, anthropic_api_key: str, tavily_api_key: str):
        self.anthropic_client = anthropic.Anthropic(api_key=anthropic_api_key)
        self.tavily_api_key = tavily_api_key
        self.model = "claude-sonnet-4-20250514"
        
        # Cost tracking
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.search_count = 0
        
    def analyze_drug(self, drug_name: str) -> Dict:
        """
        Main entry point: analyzes a drug for repurposing opportunities
        
        Returns dict with:
        - approved_indications: list of current approvals
        - case_series: list of identified opportunities
        - metadata: search statistics
        """
        logger.info(f"Starting analysis for drug: {drug_name}")
        
        # Step 1: Identify approved indications
        approved_indications = self._get_approved_indications(drug_name)
        logger.info(f"Found {len(approved_indications)} approved indications")
        
        # Step 2: Search for case series/reports
        case_series_raw = self._search_case_series(drug_name, approved_indications)
        logger.info(f"Found {len(case_series_raw)} potential case series")
        
        # Step 3: Extract structured data from each case series
        case_series_structured = []
        for idx, case in enumerate(case_series_raw, 1):
            logger.info(f"Extracting data from case {idx}/{len(case_series_raw)}")
            try:
                structured = self._extract_structured_data(drug_name, case)
                if structured:
                    case_series_structured.append(structured)
            except Exception as e:
                logger.error(f"Error extracting case {idx}: {e}")
                continue
        
        # Step 4: Enrich with epidemiology and market data
        for case in case_series_structured:
            try:
                case['market_data'] = self._get_market_data(case['disease'])
                case['soc_data'] = self._get_standard_of_care(case['disease'])
            except Exception as e:
                logger.error(f"Error enriching {case['disease']}: {e}")
                case['market_data'] = {}
                case['soc_data'] = {}
        
        # Step 5: Score opportunities
        for case in case_series_structured:
            case['scores'] = self._score_opportunity(case)
        
        # Step 6: Rank order
        case_series_structured.sort(
            key=lambda x: x['scores']['overall_priority'], 
            reverse=True
        )
        
        return {
            'drug_name': drug_name,
            'approved_indications': approved_indications,
            'case_series': case_series_structured,
            'metadata': {
                'total_input_tokens': self.total_input_tokens,
                'total_output_tokens': self.total_output_tokens,
                'search_count': self.search_count,
                'analysis_date': datetime.now().isoformat()
            }
        }
    
    def _get_approved_indications(self, drug_name: str) -> List[str]:
        """Identifies current FDA-approved indications"""
        
        # Search for FDA approval information
        search_results = self._tavily_search(
            f"{drug_name} FDA approved indications",
            max_results=5
        )
        
        # Use Claude to extract approved indications
        prompt = f"""Given the following search results about {drug_name}, extract ALL FDA-approved indications.

Search Results:
{json.dumps(search_results, indent=2)}

Return ONLY a JSON array of approved indication strings. Be comprehensive - include all mentioned indications.

Example format:
["Acute treatment of migraine with or without aura in adults", "Preventive treatment of episodic migraine in adults"]

Your response must be ONLY valid JSON, nothing else."""

        response = self.anthropic_client.messages.create(
            model=self.model,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        self._track_tokens(response.usage)
        
        try:
            # Parse JSON response
            content = response.content[0].text.strip()
            # Remove markdown code blocks if present
            if content.startswith("```"):
                content = content.split("\n", 1)[1]
                content = content.rsplit("```", 1)[0]
            return json.loads(content)
        except Exception as e:
            logger.error(f"Error parsing approved indications: {e}")
            return []
    
    def _search_case_series(self, drug_name: str, exclude_indications: List[str]) -> List[Dict]:
        """Searches for case reports and case series"""
        
        case_series = []
        
        # Multiple search strategies
        search_queries = [
            f"{drug_name} case report",
            f"{drug_name} case series",
            f"{drug_name} off-label use",
            f"{drug_name} expanded access",
            f"{drug_name} compassionate use efficacy"
        ]
        
        for query in search_queries:
            logger.info(f"Searching: {query}")
            results = self._tavily_search(query, max_results=10)
            
            # Filter for case reports/series in titles
            for result in results:
                title = result.get('title', '').lower()
                if any(term in title for term in ['case report', 'case series', 'case study']):
                    case_series.append(result)
            
            time.sleep(0.5)  # Rate limiting
        
        # Deduplicate by URL
        seen_urls = set()
        unique_cases = []
        for case in case_series:
            url = case.get('url', '')
            if url and url not in seen_urls:
                seen_urls.add(url)
                unique_cases.append(case)
        
        return unique_cases
    
    def _extract_structured_data(self, drug_name: str, case_data: Dict) -> Optional[Dict]:
        """Extracts structured data from a case series"""
        
        prompt = f"""You are analyzing a medical case report/series for drug repurposing opportunities.

Drug: {drug_name}
Case Data:
Title: {case_data.get('title', 'N/A')}
URL: {case_data.get('url', 'N/A')}
Content: {case_data.get('content', 'N/A')[:3000]}  # Limit to 3000 chars

Extract the following structured data. If information is not available, use null.

Return ONLY valid JSON in this exact format:
{{
  "source": "First Author et al., Journal Year; PMID or DOI",
  "year": 2024,
  "disease": "Specific disease/condition name",
  "n": 3,
  "evidence_level": "Case Report" or "Case Series",
  "patient_population": "Brief description of patient characteristics",
  "route_of_administration": "Oral", "IV", etc,
  "dose": "75 mg daily" or similar,
  "duration_of_followup": "6 months" or similar,
  "response_rate": "3/3 (100%)" or similar,
  "time_to_response": "Within 12 hours" or similar,
  "success_fail": "Success" or "Fail" or "Mixed",
  "effect_size_description": "Moderate: pain reduced 40%" or similar,
  "efficacy_summary": "2-3 sentence summary of efficacy findings",
  "safety_summary": "2-3 sentence summary of safety findings",
  "comparator_baseline": "What they tried before that failed",
  "durability_signal": "Did effects persist? Any relapse data?",
  "publication_venue": "Peer-reviewed journal" or "Conference abstract"
}}

CRITICAL: Return ONLY the JSON object, no other text. DO NOT include markdown code blocks."""

        try:
            response = self.anthropic_client.messages.create(
                model=self.model,
                max_tokens=2000,
                messages=[{"role": "user", "content": prompt}]
            )
            
            self._track_tokens(response.usage)
            
            # Parse response
            content = response.content[0].text.strip()
            # Clean markdown if present
            if content.startswith("```"):
                content = content.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            
            data = json.loads(content)
            
            # Add URL
            data['url'] = case_data.get('url', '')
            
            return data
            
        except Exception as e:
            logger.error(f"Error extracting structured data: {e}")
            return None
    
    def _get_market_data(self, disease: str) -> Dict:
        """Gets epidemiology and market data for a disease"""
        
        search_results = self._tavily_search(
            f"{disease} prevalence United States epidemiology",
            max_results=5
        )
        
        prompt = f"""Based on these search results, extract market data for {disease} in the United States.

Search Results:
{json.dumps(search_results, indent=2)}

Return ONLY valid JSON:
{{
  "us_prevalence_estimate": "~500,000-1M patients",
  "prevalence_source": "Citation or source",
  "unmet_need": "Yes" or "No",
  "unmet_need_description": "Brief explanation of unmet need"
}}

ONLY return the JSON, nothing else."""

        response = self.anthropic_client.messages.create(
            model=self.model,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        self._track_tokens(response.usage)
        
        try:
            content = response.content[0].text.strip()
            if content.startswith("```"):
                content = content.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            return json.loads(content)
        except:
            return {
                "us_prevalence_estimate": "Unknown",
                "prevalence_source": "Not found",
                "unmet_need": "Unknown",
                "unmet_need_description": "Insufficient data"
            }
    
    def _get_standard_of_care(self, disease: str) -> Dict:
        """Gets standard of care treatments and efficacy"""
        
        search_results = self._tavily_search(
            f"{disease} standard of care treatment efficacy",
            max_results=5
        )
        
        prompt = f"""Based on these search results, identify the top 3 standard of care treatments for {disease} and their efficacy rates.

Search Results:
{json.dumps(search_results, indent=2)}

Return ONLY valid JSON:
{{
  "top_treatments": [
    {{"drug": "Drug name", "efficacy_range": "60-70%", "notes": "Brief note"}},
    {{"drug": "Drug name", "efficacy_range": "50-60%", "notes": "Brief note"}},
    {{"drug": "Drug name", "efficacy_range": "40-50%", "notes": "Brief note"}}
  ],
  "competitive_landscape": "Brief 2-sentence summary"
}}

ONLY return the JSON, nothing else."""

        response = self.anthropic_client.messages.create(
            model=self.model,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        self._track_tokens(response.usage)
        
        try:
            content = response.content[0].text.strip()
            if content.startswith("```"):
                content = content.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            return json.loads(content)
        except:
            return {
                "top_treatments": [],
                "competitive_landscape": "Insufficient data"
            }
    
    def _score_opportunity(self, case: Dict) -> Dict:
        """Scores a repurposing opportunity"""
        
        # Clinical Signal Score (1-10)
        clinical_score = self._score_clinical_signal(case)
        
        # Evidence Quality Score (1-10)
        evidence_score = self._score_evidence_quality(case)
        
        # Market Opportunity Score (1-10)
        market_score = self._score_market_opportunity(case)
        
        # Feasibility Score (1-10)
        feasibility_score = 7  # Default for known drugs
        
        # Overall Priority (weighted)
        overall = (
            clinical_score * 0.4 +
            evidence_score * 0.2 +
            market_score * 0.2 +
            feasibility_score * 0.2
        )
        
        return {
            'clinical_signal': clinical_score,
            'evidence_quality': evidence_score,
            'market_opportunity': market_score,
            'feasibility': feasibility_score,
            'overall_priority': round(overall, 1)
        }
    
    def _score_clinical_signal(self, case: Dict) -> int:
        """Scores clinical signal strength (1-10)"""
        score = 5  # Base
        
        # Response rate
        response = case.get('response_rate', '')
        if '100%' in response:
            score += 2
        elif any(x in response for x in ['80%', '90%']):
            score += 1
        
        # Effect size
        effect = case.get('effect_size_description', '').lower()
        if 'strong' in effect or 'complete' in effect:
            score += 2
        elif 'moderate' in effect:
            score += 1
        
        # Durability
        durability = case.get('durability_signal', '').lower()
        if any(x in durability for x in ['months', 'sustained', 'long-term']):
            score += 1
        
        return min(10, max(1, score))
    
    def _score_evidence_quality(self, case: Dict) -> int:
        """Scores evidence quality (1-10)"""
        score = 5  # Base
        
        # Sample size
        n = case.get('n', 1)
        if n >= 10:
            score += 2
        elif n >= 3:
            score += 1
        else:
            score -= 1
        
        # Publication venue
        venue = case.get('publication_venue', '').lower()
        if 'peer-reviewed' in venue:
            score += 1
        elif 'conference' in venue:
            score -= 1
        
        # Follow-up duration
        duration = case.get('duration_of_followup', '').lower()
        if 'month' in duration or 'year' in duration:
            score += 1
        
        return min(10, max(1, score))
    
    def _score_market_opportunity(self, case: Dict) -> int:
        """Scores market opportunity (1-10)"""
        score = 5  # Base
        
        market_data = case.get('market_data', {})
        prevalence = market_data.get('us_prevalence_estimate', '').lower()
        
        # Estimate patient population
        if 'million' in prevalence:
            score += 3
        elif '000' in prevalence or 'hundred thousand' in prevalence:
            score += 1
        
        # Unmet need
        unmet = market_data.get('unmet_need', '').lower()
        if unmet == 'yes':
            score += 2
        
        return min(10, max(1, score))
    
    def _tavily_search(self, query: str, max_results: int = 5) -> List[Dict]:
        """Performs Tavily search"""
        self.search_count += 1
        
        try:
            response = requests.post(
                "https://api.tavily.com/search",
                json={
                    "api_key": self.tavily_api_key,
                    "query": query,
                    "max_results": max_results,
                    "search_depth": "advanced"
                },
                timeout=30
            )
            response.raise_for_status()
            results = response.json()
            return results.get('results', [])
        except Exception as e:
            logger.error(f"Tavily search error: {e}")
            return []
    
    def _track_tokens(self, usage):
        """Tracks token usage for cost estimation"""
        self.total_input_tokens += usage.input_tokens
        self.total_output_tokens += usage.output_tokens
    
    def export_to_excel(self, analysis_results: Dict, output_path: str):
        """Exports results to Excel file"""
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Repurposing Opportunities"
        
        # Headers
        headers = [
            "Source", "Drug", "Year", "Disease", "N", "Evidence Level",
            "Patient Population", "RoA", "Dose", "Follow-up", "Response Rate",
            "Time to Response", "Success/Fail", "Effect Size", "Efficacy",
            "Safety", "Comparator", "Durability", "Publication Venue",
            "US Prevalence", "Current SOC", "Unmet Need", "Competitive Landscape",
            "Clinical Score", "Evidence Score", "Market Score", "Feasibility Score",
            "Overall Priority"
        ]
        
        # Format headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Data rows
        for row_idx, case in enumerate(analysis_results['case_series'], 2):
            scores = case.get('scores', {})
            market = case.get('market_data', {})
            soc = case.get('soc_data', {})
            
            # Format SOC treatments
            soc_text = ""
            for treatment in soc.get('top_treatments', [])[:3]:
                soc_text += f"{treatment.get('drug', 'N/A')}: {treatment.get('efficacy_range', 'N/A')}; "
            
            row_data = [
                case.get('source', ''),
                analysis_results['drug_name'],
                case.get('year', ''),
                case.get('disease', ''),
                case.get('n', ''),
                case.get('evidence_level', ''),
                case.get('patient_population', ''),
                case.get('route_of_administration', ''),
                case.get('dose', ''),
                case.get('duration_of_followup', ''),
                case.get('response_rate', ''),
                case.get('time_to_response', ''),
                case.get('success_fail', ''),
                case.get('effect_size_description', ''),
                case.get('efficacy_summary', ''),
                case.get('safety_summary', ''),
                case.get('comparator_baseline', ''),
                case.get('durability_signal', ''),
                case.get('publication_venue', ''),
                market.get('us_prevalence_estimate', ''),
                soc_text,
                market.get('unmet_need_description', ''),
                soc.get('competitive_landscape', ''),
                scores.get('clinical_signal', ''),
                scores.get('evidence_quality', ''),
                scores.get('market_opportunity', ''),
                scores.get('feasibility', ''),
                scores.get('overall_priority', '')
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            sheet.row_dimensions[row_idx].height = 60
        
        # Freeze header
        sheet.freeze_panes = "A2"
        
        # Add metadata sheet
        meta_sheet = wb.create_sheet("Metadata")
        meta_sheet.column_dimensions['A'].width = 30
        meta_sheet.column_dimensions['B'].width = 50
        
        metadata = [
            ("Drug Analyzed", analysis_results['drug_name']),
            ("Analysis Date", analysis_results['metadata']['analysis_date']),
            ("Approved Indications", "; ".join(analysis_results['approved_indications'])),
            ("Total Opportunities Found", len(analysis_results['case_series'])),
            ("Total API Calls", self.search_count),
            ("Total Input Tokens", self.total_input_tokens),
            ("Total Output Tokens", self.total_output_tokens),
            ("Estimated Cost", f"${(self.total_input_tokens * 0.003 / 1000 + self.total_output_tokens * 0.015 / 1000):.2f}")
        ]
        
        for idx, (key, value) in enumerate(metadata, 1):
            meta_sheet.cell(row=idx, column=1).value = key
            meta_sheet.cell(row=idx, column=1).font = Font(bold=True)
            meta_sheet.cell(row=idx, column=2).value = str(value)
        
        wb.save(output_path)
        logger.info(f"Results exported to {output_path}")


def main():
    """Example usage"""
    import os
    
    # Get API keys from environment
    ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
    TAVILY_API_KEY = os.getenv("TAVILY_API_KEY")
    
    if not ANTHROPIC_API_KEY or not TAVILY_API_KEY:
        print("Error: Please set ANTHROPIC_API_KEY and TAVILY_API_KEY environment variables")
        return
    
    # Initialize agent
    agent = DrugRepurposingAgent(
        anthropic_api_key=ANTHROPIC_API_KEY,
        tavily_api_key=TAVILY_API_KEY
    )
    
    # Analyze drug
    drug_name = "rimegepant"
    print(f"Analyzing {drug_name}...")
    
    results = agent.analyze_drug(drug_name)
    
    # Export results
    output_file = f"{drug_name}_repurposing_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
    agent.export_to_excel(results, output_file)
    
    print(f"\nAnalysis complete!")
    print(f"Found {len(results['case_series'])} repurposing opportunities")
    print(f"Total cost: ${(agent.total_input_tokens * 0.003 / 1000 + agent.total_output_tokens * 0.015 / 1000):.2f}")
    print(f"Results saved to: {output_file}")
    
    # Print top 3 opportunities
    print("\nTop 3 Opportunities:")
    for idx, case in enumerate(results['case_series'][:3], 1):
        print(f"\n{idx}. {case['disease']}")
        print(f"   Priority Score: {case['scores']['overall_priority']}/10")
        print(f"   N={case['n']}, Response: {case['response_rate']}")
        print(f"   {case['efficacy_summary'][:100]}...")


if __name__ == "__main__":
    main()
